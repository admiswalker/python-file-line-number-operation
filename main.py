import io
import pandas as pd
import openpyxl

def main():
    fp = open('./example.csv', 'r', encoding='UTF-8')
    s = fp.read()
    fp.close()
    
    print(s)
    
    #df = pd.read_csv('./example.csv', comment='#')
    df = pd.read_csv(io.StringIO(s), comment='#')
    print(df)

def test():
    df = pd.read_excel('./example.xlsx', comment='#')
    print(df)

def test_excel_to_csv():
    read_file = pd.read_excel('example_with_comment1.xlsx')
    read_file.to_csv('./example_.csv')

def test_csv_to_excel():
    read_file = pd.read_csv('./example.csv')
    read_file.to_excel('./example.xlsx')

def read_excel_01():
    wb = openpyxl.load_workbook('./example.xlsx')
    #l_sheet_name = wb.sheetnames
    #ws1_name = l_sheet_name[0]
    #ws1 = wb[ws1_name]

    for ws in wb:
        for row in ws.rows:
            for cell in row:
                print(cell.value, end='')
                print(', ', end='')
            print()

def read_excel_02():
    wb = openpyxl.load_workbook('./example_with_comment1.xlsx')
    #l_sheet_name = wb.sheetnames
    #ws1_name = l_sheet_name[0]
    #ws1 = wb[ws1_name]

    for ws in wb:
        for row in ws.rows:
            for cell in row:
                print(cell.value, end='')
                print(', ', end='')
            print()


#main()
#test_csv_to_excel()
#test()
#test_excel_to_csv()

#read_excel()
read_excel_02()



# memo
#
# - Pandas - read csv stored as string in memory to data frame
#   https://stackoverflow.com/questions/58288236/pandas-read-csv-stored-as-string-in-memory-to-data-frame
# - openpyxl による Excelファイル操作方法のまとめ
#   https://gammasoft.jp/support/how-to-use-openpyxl-for-excel-file/#get-sheet

